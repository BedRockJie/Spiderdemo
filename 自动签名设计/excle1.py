from openpyxl import Workbook
wb = Workbook()
#print(wb.get_sheet_names()) # 提供一个默认名叫Sheet的表，office2016下新建提供默认Sheet1
sheet = wb.active
# 直接赋值就可以改工作表的名称
sheet.title = 'Sheet1'
# 新建一个工作表，可以指定索引，适当安排其在工作簿中的位置
# wb.create_sheet('Data', index=1) # 被安排到第二个工作表，index=0就是第一个位置
# # 删除某个工作表
# wb.remove(sheet)
headDate = ["职位名称","地点","时间","行业","招聘时间","人数","顾问"]
sheet['A1'] = 'good'
for i in range(len(headDate)):
    #sheet["A{}".format(i)] = headDate[i]
    sheet.cell(row=1,column=i+1).value = headDate[i]
wb.save(r'example.xlsx')