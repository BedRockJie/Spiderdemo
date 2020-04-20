import urllib.request
import re
from openpyxl import Workbook

def getdate():
    for i in range(26700,26716):
        url = 'http://www.risfond.com/case/fmcg/{}'.format(i)
        html = urllib.request.urlopen(url).read().decode('utf-8')
        # item = ge
        page_list = re.findall('<div class="sc_d_c">.*?<span class="sc_d_con">(.*?)</span></div>',html)
        # print(page_list)
        yield page_list

def excle_write(date):

    #item = list(date)
   # print(item)
    wb = Workbook()
    sheet = wb.active
    sheet.tittle = "数据表格"
    headDate = ["职位名称","地点","时间","行业","招聘时间","人数","顾问"]
    # for head in headDate:
    #     sheet[A]
    for i in range(len(headDate)):
        sheet.cell(row=1,column=i+1).value = headDate[i]
    for i in range(10):
        for j in range(7):
            sheet.cell(row=i+2,column=j+1).value = date[i][j]
    wb.save(r"example.xlsx")

try:
    items = getdate()
    #print(items)
except urllib.error.HTTPError:
    print("请求超时！")
else:
    date = list(items)
    print(date)
    excle_write(date)
  #  print(type(date))
  #  print(date)

